#!/usr/bin/env python3
"""Ingest WB SAEU April 2026 data into Indica series JSONs.

Source report: https://www.worldbank.org/en/region/sar/publication/south-asia-economic-update
"""

import json, os
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

WB_DIR = Path.home() / "Documents/Data/WB/SAEU-April-2026-Charts"
OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "series"
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
SOURCE_URL = "https://www.worldbank.org/en/region/sar/publication/south-asia-economic-update"

os.makedirs(OUT_DIR, exist_ok=True)

GEO_MULTI = {"type": "multi-country", "id": "SAR", "name": "South Asia"}


def load_wb(name):
    for ext in [".xlsx", ".xlsm"]:
        p = WB_DIR / f"{name}{ext}"
        if p.exists():
            return openpyxl.load_workbook(p, data_only=True)
    raise FileNotFoundError(name)


def write_json(data):
    fp = OUT_DIR / f"{data['indicatorId']}.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)
    print(f"  {fp.name}")


def series(iid, title, unit, obs, **meta):
    return {
        "schemaVersion": 1, "artifactType": "series",
        "indicatorId": iid, "title": title,
        "sourceId": "worldbank", "sourceIndicatorId": "SAEU April 2026",
        "sourceUrl": SOURCE_URL, "unit": unit,
        "frequency": "annual", "geography": GEO_MULTI,
        "dimensions": [], "fetchedAt": NOW,
        "observations": obs, "metadata": meta,
    }


def table(iid, title, unit, rows, **meta):
    return {
        "schemaVersion": 1, "artifactType": "table",
        "indicatorId": iid, "title": title,
        "sourceId": "worldbank", "sourceIndicatorId": "SAEU April 2026",
        "sourceUrl": SOURCE_URL, "unit": unit,
        "geography": GEO_MULTI,
        "dimensions": list(rows[0].keys()) if rows else [],
        "fetchedAt": NOW, "rows": rows, "metadata": meta,
    }


def strip_nones(vals):
    while vals and vals[0] is None:
        vals.pop(0)
    return vals


def iter_data_rows(ws, start_row=1):
    """Yield stripped, non-empty data rows from a worksheet."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
        vals = strip_nones(list(row))
        if not vals or not any(v is not None and str(v).strip() for v in vals):
            continue
        first = str(vals[0]).strip() if vals[0] else ""
        if first.startswith(("Sources:", "Note:", "Return", "Source:")):
            continue
        yield vals


# ═══════════════════════════════════════════════════════
wb_ch2 = load_wb("SAEU-April-2026-Chapter-2-charts")
wb_b11 = load_wb("SAEU-April-2026-Box-1-1-charts")


# ── Chart 1: SA protective IP measures ─────────────────
def ingest_sa_protective():
    # Fig 2.1.E: 2022-25  and  Fig 2.1.F: 2016-19
    # Row format after strip: [country(RHS), value, None, emde_p25, emde_median, emde_iqr]
    # Or for non-IND: [country, None, value, emde_p25, emde_median, emde_iqr]
    # Manual extraction for reliability:

    # Known values from WB SAEU April 2026
    data_2225 = {"India": 239.3, "Bangladesh": 12, "Sri Lanka": 9, "Nepal": 2.8, "Bhutan": 0.3}
    data_1619 = {"India": 124.8, "Sri Lanka": 14.5, "Bangladesh": 2.8, "Nepal": 1.3, "Maldives": 0.8}

    for c in ["India", "Bangladesh", "Sri Lanka", "Nepal", "Bhutan"]:
        v1619 = data_1619.get(c)
        v2225 = data_2225.get(c)
        if v1619 is not None:
            write_json(table(
                f"ipp.protective_measures.{c.lower().replace(' ', '_')}_1619",
                f"Protective industrial policy measures: {c}, 2016-19 average",
                "measures per year",
                [{"country": c, "period": "2016-19", "measures": v1619,
                  "emde_median": 1.0, "emde_p25": 0.5}],
                note=f"Source: GTA via WB SAEU April 2026 Fig 2.1.F."
            ))
        if v2225 is not None:
            write_json(table(
                f"ipp.protective_measures.{c.lower().replace(' ', '_')}_2225",
                f"Protective industrial policy measures: {c}, 2022-25 average",
                "measures per year",
                [{"country": c, "period": "2022-25", "measures": v2225,
                  "emde_median": 2.8, "emde_p25": 0.8}],
                note=f"Source: GTA via WB SAEU April 2026 Fig 2.1.E."
            ))

    # Combined table
    rows = []
    for c, v1619 in data_1619.items():
        if v1619 is not None:
            rows.append({"country": c, "period": "2016-19", "measures": v1619})
    for c, v2225 in data_2225.items():
        if v2225 is not None:
            rows.append({"country": c, "period": "2022-25", "measures": v2225})
    write_json(table(
        "ipp.sa_protective_measures",
        "South Asia: protective industrial policy measures per year",
        "measures per year",
        rows,
        note="EMDE median: 1.0 (2016-19), 2.8 (2022-25). Source: GTA via WB SAEU April 2026."
    ))


# ── Chart 2: Global IP wave ────────────────────────────
def ingest_global_wave():
    ws = wb_ch2["2.1.A"]
    obs_total, obs_prot = [], []
    for vals in iter_data_rows(ws, 3):
        if len(vals) >= 2 and isinstance(vals[0], (int, float)):
            yr = str(int(vals[0]))
            obs_total.append({"date": yr, "value": float(vals[1]) if vals[1] else None})
            if len(vals) >= 3:
                obs_prot.append({"date": yr, "value": float(vals[2]) if vals[2] else None})
    write_json(series("ipp.world_total_measures", "World: total new industrial policy measures",
                      "number of measures", obs_total,
                      note="Source: GTA via WB SAEU April 2026 Fig 2.1.A."))
    write_json(series("ipp.world_protective_measures", "World: new protective industrial policy measures",
                      "number of measures", obs_prot,
                      note="Protective = discriminates against foreign businesses. Source: GTA via WB SAEU April 2026."))


# ── Chart 3: Global ranking ────────────────────────────
def ingest_global_ranking():
    ws = wb_ch2["2.1.D"]
    rows = []
    for vals in iter_data_rows(ws, 3):
        country = str(vals[0]).strip() if vals[0] else ""
        v2225 = vals[1] if len(vals) > 1 else None
        v1619 = vals[2] if len(vals) > 2 else None
        if country and v2225 is not None:
            rows.append({"country": country,
                         "measures_2022_25": float(v2225),
                         "measures_2016_19": float(v1619) if v1619 else None})
    write_json(table("ipp.top_countries_protective",
                     "Top countries by protective industrial policy measures, 2022-25 average",
                     "measures per year", rows,
                     note="Top 10 countries globally. Source: GTA via WB SAEU April 2026 Fig 2.1.D."))


# ── Chart 4: Instrument mix ────────────────────────────
def ingest_instrument_mix():
    ws = wb_ch2["2.3.B"]
    # Header: BGD IND LKA NPL at positions 0-3 after strip
    instruments = []
    for vals in iter_data_rows(ws, 2):
        inst = str(vals[0]).strip() if vals[0] else ""
        if inst in ("Imports", "Exports", "Procurement", "Subsidies", "Other"):
            instruments.append({
                "instrument": inst,
                "Bangladesh": float(vals[1]) if len(vals) > 1 and vals[1] is not None else 0,
                "India": float(vals[2]) if len(vals) > 2 and vals[2] is not None else 0,
                "Sri Lanka": float(vals[3]) if len(vals) > 3 and vals[3] is not None else 0,
                "Nepal": float(vals[4]) if len(vals) > 4 and vals[4] is not None else 0,
            })
    write_json(table("ipp.instrument_mix",
                     "South Asia: protective industrial policy measures by instrument, 2022-25",
                     "percent of protective measures", instruments,
                     note="Source: GTA via WB SAEU April 2026 Fig 2.3.B."))


# ── Chart 5: Sectors targeted ──────────────────────────
def ingest_sectors():
    rows_out = []
    for sheet, period in [("2.5C", "2022-25"), ("2.5D", "2016-19")]:
        ws = wb_ch2[sheet]
        for vals in iter_data_rows(ws, 3):
            # Cols: [rank, ISIC, SAR_share, OtherEMDE_share, Label]
            sector = str(vals[-1]).strip() if vals and vals[-1] else None
            sar = vals[2] if len(vals) > 2 else None
            emde = vals[3] if len(vals) > 3 else None
            if sector and sar is not None:
                rows_out.append({
                    "sector": sector, "period": period,
                    "sar_share_pct": float(sar),
                    "other_emde_share_pct": float(emde) if emde else None,
                })
    write_json(table("ipp.sectors_targeted",
                     "South Asia: manufacturing sectors targeted by protective industrial policy",
                     "percent of protective IP measures", rows_out,
                     note="Source: GTA via WB SAEU April 2026 Fig 2.5.C-D."))


# ── Chart 6: Tariff wall ───────────────────────────────
def ingest_tariff_wall():
    write_json(table("ipp.average_import_duties",
                     "Average import duty, latest available",
                     "percent",
                     [{"country": "India", "rate": 15.8, "type": "Simple avg MFN"},
                      {"country": "Sri Lanka", "rate": 19.0, "type": "Includes para-tariffs"},
                      {"country": "Bangladesh", "rate": 14.0, "type": "GDP-weighted SAR others"},
                      {"country": "EMDE 25th pctile", "rate": 5.6, "type": ""},
                      {"country": "EMDE median", "rate": 8.5, "type": ""},
                      {"country": "EMDE 75th pctile", "rate": 12.0, "type": ""}],
                     note="Source: WTO Analytical DB via WB SAEU April 2026 Fig B1.1.1.C."))


# ── Chart 7: Policy drivers ────────────────────────────
def ingest_policy_drivers():
    # Manually curated from regression tables 2.6 and 2.7
    write_json(table("ipp.policy_drivers_india",
                     "India: sector characteristics predicting more protective IP, 2022-25",
                     "regression coefficient",
                     [
                         {"factor": "Import share (+1 pp)", "coefficient": 0.47, "ci": 0.14,
                          "note": "Higher import penetration strongly predicts more protection"},
                         {"factor": "Export share (+1 pp)", "coefficient": 0.39, "ci": 0.13,
                          "note": "Export-oriented sectors also get more IP (dual-purpose)"},
                         {"factor": "Hourly wage (+10%) — mfg", "coefficient": 0.24, "ci": 0.15,
                          "note": "In manufacturing, higher-wage sectors get more protection"},
                         {"factor": "Firm employment (+10%)", "coefficient": 0.21, "ci": 0.12,
                          "note": "Sectors with larger firms get more protection"},
                         {"factor": "Firm productivity (+10%)", "coefficient": 0.15, "ci": None,
                          "note": "More productive sectors get more protection (India)"},
                     ],
                     note="Regression of sector share of new IP on lagged characteristics. WB SAEU April 2026 Figs 2.6-2.7."))


# ── Chart 8: Trade effects ─────────────────────────────
def ingest_trade_effects():
    # 2.8A: imports after trade defense
    ws = wb_ch2["2.8A"]
    obs_imp = []
    for vals in iter_data_rows(ws, 3):
        if len(vals) >= 4 and isinstance(vals[0], (int, float)):
            obs_imp.append({"date": str(int(vals[0])),
                            "value": float(vals[1]),
                            "lower": float(vals[2]) if vals[2] else None,
                            "upper": float(vals[3]) if vals[3] else None})
    write_json(series("ipp.trade_defense_imports",
                      "SA: cumulative change in imports after trade defense instruments",
                      "percent", obs_imp,
                      note="Local projection. t=0 is policy start. 90% CI. WB SAEU April 2026 Fig 2.8.A."))

    # 2.8C: exports after export incentives
    ws = wb_ch2["2.8C"]
    obs_exp = []
    for vals in iter_data_rows(ws, 3):
        if len(vals) >= 4 and isinstance(vals[0], (int, float)):
            obs_exp.append({"date": str(int(vals[0])),
                            "value": float(vals[1]),
                            "lower": float(vals[2]) if vals[2] else None,
                            "upper": float(vals[3]) if vals[3] else None})
    write_json(series("ipp.export_incentive_exports",
                      "SA: cumulative change in exports after export incentives",
                      "percent", obs_exp,
                      note="Local projection. t=0 is policy start. 90% CI. WB SAEU April 2026 Fig 2.8.C."))


# ── Chart 9: Employment by IP quartile ─────────────────
def ingest_employment_effects():
    rows_out = []
    for sheet, sector_type in [("2.5E", "Non-agriculture"), ("2.5F", "Manufacturing")]:
        ws = wb_ch2[sheet]
        for vals in iter_data_rows(ws, 3):
            label = str(vals[0]).strip() if vals[0] else ""
            if label in ("Lowest quartile", "Middle quartiles"):
                for i, c in enumerate(["Bangladesh", "India", "Sri Lanka"]):
                    val = vals[i + 1] if len(vals) > i + 1 else None
                    if val is not None:
                        rows_out.append({
                            "country": c, "sector_type": sector_type,
                            "ip_quartile": label, "annual_emp_growth_pct": float(val),
                        })
    write_json(table("ipp.employment_by_ip_quartile",
                     "SA: annual employment growth by industrial policy intensity quartile",
                     "percent per year", rows_out,
                     note="Sectors grouped by quartile of IP measures received. WB SAEU April 2026 Fig 2.5.E-F."))


# ── Chart 10: FTA household effects ────────────────────
def ingest_fta_household():
    # B1.1.2.E consumption effects
    ws = wb_b11["B1.1.2.E"]
    consumption = []
    for vals in iter_data_rows(ws, 2):
        area = str(vals[0]).strip() if vals[0] else ""
        if area in ("Rural", "Urban"):
            for i, q in enumerate(["Q1", "Q2", "Q3", "Q4", "Q5"]):
                val = vals[i + 1] if len(vals) > i + 1 else None
                if val is not None:
                    consumption.append({"area": area, "quintile": q, "consumption_effect_pct": float(val)})

    # B1.1.2.F income effects
    ws = wb_b11["B1.1.2.F"]
    income = []
    for vals in iter_data_rows(ws, 2):
        area = str(vals[0]).strip() if vals[0] else ""
        if area in ("Rural", "Urban"):
            for i, q in enumerate(["Q1", "Q2", "Q3", "Q4", "Q5"]):
                val = vals[i + 1] if len(vals) > i + 1 else None
                if val is not None:
                    income.append({"area": area, "quintile": q, "real_income_effect_pct": float(val)})

    write_json(table("ipp.fta_consumption_effects",
                     "India: consumption effects of EU/UK FTA tariff cuts on households",
                     "percent change", consumption,
                     note="First-order effects. WB SAEU April 2026 Fig B1.1.2.E."))
    write_json(table("ipp.fta_income_effects",
                     "India: real income effects of EU/UK FTA tariff cuts on households",
                     "percent change", income,
                     note="First-order effects. WB SAEU April 2026 Fig B1.1.2.F."))


# ── Chart 11: India tariff cuts vs RCA ─────────────────
def ingest_india_tariff_cuts():
    # B1.1.3.A: output tariff cuts
    ws = wb_b11["B1.1.3.A"]
    output_cuts = {}
    for vals in iter_data_rows(ws, 2):
        if len(vals) >= 5 and str(vals[0]).strip() == "India":
            sector = str(vals[4]).strip()
            output_cuts[sector] = {
                "log_rca": float(vals[1]) if vals[1] else None,
                "tariff_cut_output_pp": float(vals[2]) if vals[2] else None,
                "export_share_pct": float(vals[3]) if vals[3] else None,
            }

    # B1.1.3.C: effective tariff cuts
    ws = wb_b11["B1.1.3.C"]
    effective = {}
    for vals in iter_data_rows(ws, 2):
        if len(vals) >= 5 and str(vals[0]).strip() == "India":
            sector = str(vals[4]).strip()
            effective[sector] = float(vals[2]) if vals[2] else None

    rows = []
    for sector, data in output_cuts.items():
        r = {"sector": sector, **data, "effective_tariff_cut_pp": effective.get(sector)}
        rows.append(r)

    write_json(table("ipp.india_tariff_cuts_rca",
                     "India: FTA tariff-cut commitments and comparative advantage by sector",
                     "percentage points / log RCA", rows,
                     note="Tariff cuts from India-EU/UK FTAs. Log RCA > 0 = comparative advantage. WB SAEU April 2026 Fig B1.1.3."))


# ── Charts 12-13: Fiscal & capacity ────────────────────
def ingest_fiscal_and_capacity():
    # Tax revenue 2.9.A
    write_json(table("ipp.tax_revenue",
                     "South Asia: tax revenue, 2019-23 average",
                     "percent of GDP",
                     [{"country": "Maldives", "tax_pct_gdp": 20.1, "emde_avg": 20.4},
                      {"country": "Nepal", "tax_pct_gdp": 18.6, "emde_avg": 20.4},
                      {"country": "India", "tax_pct_gdp": 18.0, "emde_avg": 20.4},
                      {"country": "Bhutan", "tax_pct_gdp": 14.0, "emde_avg": 20.4},
                      {"country": "Sri Lanka", "tax_pct_gdp": 10.0, "emde_avg": 20.4},
                      {"country": "Bangladesh", "tax_pct_gdp": 8.5, "emde_avg": 20.4}],
                     note="Includes social security contributions, excludes grants. WB SAEU April 2026 Fig 2.9.A."))

    # Customs 2.10.B
    write_json(table("ipp.customs_clearance_days",
                     "South Asia: days to clear customs for medium-size firms",
                     "days",
                     [{"country": "India", "import_days": 16.3, "export_days": 19.9, "emde_median_import": 5.1, "emde_median_export": 5.3},
                      {"country": "Bangladesh", "import_days": 8.9, "export_days": 8.9, "emde_median_import": 5.1, "emde_median_export": 5.3},
                      {"country": "Bhutan", "import_days": 5.5, "export_days": 6.0, "emde_median_import": 5.1, "emde_median_export": 5.3},
                      {"country": "Nepal", "import_days": 4.0, "export_days": 4.5, "emde_median_import": 5.1, "emde_median_export": 5.3}],
                     note="Medium-size firms (20-99 employees). WB SAEU April 2026 Fig 2.10.B."))

    # Govt effectiveness 2.9.D
    write_json(table("ipp.govt_effectiveness",
                     "South Asia: government effectiveness, 2024",
                     "percentile rank (0-100)",
                     [{"country": "Bhutan", "rank": 69.3, "emde_median": 37.3},
                      {"country": "India", "rank": 67.9, "emde_median": 37.3},
                      {"country": "Maldives", "rank": 45.0, "emde_median": 37.3},
                      {"country": "Sri Lanka", "rank": 40.0, "emde_median": 37.3},
                      {"country": "Bangladesh", "rank": 30.0, "emde_median": 37.3},
                      {"country": "Nepal", "rank": 20.0, "emde_median": 37.3}],
                     note="Worldwide Governance Indicators. AE median: 87.7. WB SAEU April 2026 Fig 2.9.D."))

    # FTA coverage B1.1.2.A
    write_json(table("ipp.fta_coverage",
                     "Share of global output covered by free trade agreements",
                     "percent of global GDP",
                     [{"country": "Mexico", "current_pct": 60.8, "concluded_pct": 0.0, "emde_median": 18.6},
                      {"country": "Vietnam", "current_pct": 58.4, "concluded_pct": 1.0, "emde_median": 18.6},
                      {"country": "China", "current_pct": 30.3, "concluded_pct": 2.3, "emde_median": 18.6},
                      {"country": "Turkey", "current_pct": 26.8, "concluded_pct": 0.3, "emde_median": 18.6},
                      {"country": "India", "current_pct": 15.5, "concluded_pct": 17.5, "emde_median": 18.6},
                      {"country": "Brazil", "current_pct": 3.7, "concluded_pct": 18.0, "emde_median": 18.6}],
                     note="Current FTAs + negotiations concluded. India's EU+UK deals double coverage. WB SAEU April 2026 Fig B1.1.2.A."))


# ═══════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Ingesting WB SAEU April 2026...")
    ingest_sa_protective()
    ingest_global_wave()
    ingest_global_ranking()
    ingest_instrument_mix()
    ingest_sectors()
    ingest_tariff_wall()
    ingest_policy_drivers()
    ingest_trade_effects()
    ingest_employment_effects()
    ingest_fta_household()
    ingest_india_tariff_cuts()
    ingest_fiscal_and_capacity()
    print("Done.")
