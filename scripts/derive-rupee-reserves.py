#!/usr/bin/env python3
"""Splice RBI monthly total reserves into the long rupee reserves artifact.

The original article series used a long FRED-plus-derived history. RBI's
monthly workbook is the better source for the recent total because it gives the
component values in US dollars. Keep the older history before RBI's workbook
starts, then replace the overlapping months with RBI totals.
"""

from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ARTIFACT = ROOT / "data/series/rupee-derived.derived.IN.reserves.total_usd.monthly.json"
RBI_SOURCE_URL = "https://www.rbi.org.in/Scripts/BS_ViewBulletin.aspx"
MONTHS = {
    "JAN": "01",
    "FEB": "02",
    "MAR": "03",
    "APR": "04",
    "MAY": "05",
    "JUN": "06",
    "JUL": "07",
    "AUG": "08",
    "SEP": "09",
    "OCT": "10",
    "NOV": "11",
    "DEC": "12",
}


def find_workbook() -> Path:
    env = os.environ.get("RBI_RESERVES_XLSX")
    candidates = [Path(env)] if env else []
    candidates += sorted(
        Path("/home/bhuvanesh.r/Downloads").glob("Foreign Exchange Reserves*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for path in candidates:
        if path and path.exists():
            return path
    raise FileNotFoundError("Set RBI_RESERVES_XLSX or place Foreign Exchange Reserves*.xlsx in Downloads")


def read_rbi_monthly(path: Path) -> dict[str, float]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    observations: dict[str, float] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        year, month = row[1], row[2]
        if not year or not month:
            continue
        month_code = MONTHS.get(str(month).strip().upper()[:3])
        if not month_code:
            continue
        usd_components = [row[4], row[6], row[8], row[11]]
        if any(value is None or str(value).strip() in {"", "-"} for value in usd_components):
            continue
        date = f"{int(year):04d}-{month_code}"
        observations[date] = round(sum(float(value) for value in usd_components) / 1000, 3)
    if not observations:
        raise ValueError(f"No RBI reserves observations found in {path}")
    return observations


def main() -> None:
    workbook = find_workbook()
    rbi = read_rbi_monthly(workbook)
    first_rbi_month = min(rbi)

    artifact = json.loads(ARTIFACT.read_text())
    old = artifact["observations"]
    merged = {obs["date"]: obs["value"] for obs in old if obs["date"] < first_rbi_month}
    merged.update(rbi)

    artifact["sourceUrl"] = RBI_SOURCE_URL
    artifact["fetchedAt"] = datetime.now(timezone.utc).isoformat()
    artifact["observations"] = [
        {"date": date, "value": merged[date]} for date in sorted(merged)
    ]
    artifact.setdefault("metadata", {})
    artifact["metadata"].update(
        {
            "derivation": "Long reserves history preserved before the RBI workbook starts; overlapping and latest months use RBI monthly component totals in US dollars, summed and converted from US$ millions to US$ billions.",
            "rbiWorkbook": str(workbook),
            "rbiFirstMonth": first_rbi_month,
            "rbiLastMonth": max(rbi),
        }
    )
    ARTIFACT.write_text(json.dumps(artifact, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {ARTIFACT} with {len(artifact['observations'])} observations")
    print(f"RBI splice: {first_rbi_month} -> {max(rbi)}, latest {rbi[max(rbi)]} US$ billion")


if __name__ == "__main__":
    main()
