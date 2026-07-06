#!/usr/bin/env python3
"""Ingest state GSDP / NSDP / per-capita income from the RBI Handbook of
Statistics on Indian States 2024-25 (Tables 19, 21, 23).

These unlock the per-GSDP and per-capita normalisation the demographic-finances
flagship deferred (e-STATES has no GSDP column). Files were downloaded from
rbidocs.rbi.org.in via a real browser (the host is Akamai bot-walled to plain
HTTP clients).

  Table 21: Gross State Domestic Product, current prices (Rs Lakh), 2011-12 series
  Table 23: Net State Domestic Product, current prices (Rs Lakh)
  Table 19: Per Capita Net State Domestic Product, current prices (Rs)

Each workbook splits years across two sheets ((i) 2011-12..2016-17,
(ii) 2017-18..2024-25); we merge them. State population is derived as
NSDP_total / per-capita-NSDP (both from the same Handbook, internally consistent).

Source: https://www.rbi.org.in/Scripts/AnnualPublications.aspx?head=Handbook+of+Statistics+on+Indian+States
"""

import json
import os
import re

import openpyxl

DL = os.path.expanduser("~/Downloads")
FILES = {
    "gsdp": "21T_11122025D994949B48C44B68B4465FBB9ADDFF3D.XLSX",      # Rs Lakh
    "pc_nsdp": "19T_11122025B8CC230E4A34431999B4D6A107707BCA.XLSX",   # Rs
    "nsdp": "23T_11122025142A7C614BA04533BA1EE9EC9BCA7DBB.XLSX",      # Rs Lakh
}
OUT_DIR = "data/snapshots/estates"
OUT = os.path.join(OUT_DIR, "rbi_handbook_states.json")
SOURCE_URL = ("https://www.rbi.org.in/Scripts/AnnualPublications.aspx"
              "?head=Handbook+of+Statistics+on+Indian+States")

YEAR_RE = re.compile(r"^\d{4}-\d{2}$")


def canon(name):
    n = name.strip().rstrip("*").strip()
    n = n.replace(" & ", " and ")  # Jammu & Kashmir -> Jammu and Kashmir
    return n


def parse_table(path):
    """Return {canonical_state: {year_start_str: float}} merged across sheets."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sh in wb.sheetnames:
        rows = list(wb[sh].iter_rows(values_only=True))
        hdr_idx = None
        for i, r in enumerate(rows[:8]):
            if any(isinstance(c, str) and YEAR_RE.match(c.strip()) for c in r if c):
                hdr_idx = i
                break
        if hdr_idx is None:
            continue
        col_year = {ci: c.strip().split("-")[0]
                    for ci, c in enumerate(rows[hdr_idx])
                    if isinstance(c, str) and YEAR_RE.match(c.strip())}
        for r in rows[hdr_idx + 1:]:
            if len(r) < 2 or not (isinstance(r[1], str) and r[1].strip()):
                continue
            d = out.setdefault(canon(r[1]), {})
            for ci, yr in col_year.items():
                v = r[ci] if ci < len(r) else None
                if v in (None, "-", ""):
                    continue
                try:
                    d[yr] = float(v)
                except (TypeError, ValueError):
                    pass
    return out


def main():
    gsdp = parse_table(os.path.join(DL, FILES["gsdp"]))     # Rs Lakh
    pc = parse_table(os.path.join(DL, FILES["pc_nsdp"]))    # Rs
    nsdp = parse_table(os.path.join(DL, FILES["nsdp"]))     # Rs Lakh

    states = sorted(set(gsdp) | set(pc) | set(nsdp))
    payload = {}
    for s in states:
        g = gsdp.get(s, {})
        p = pc.get(s, {})
        n = nsdp.get(s, {})
        pop = {}
        for yr in n:
            if p.get(yr):
                # NSDP in Rs Lakh -> Rs; divide by per-capita Rs -> persons
                pop[yr] = round(n[yr] * 1e5 / p[yr])
        payload[s] = {
            "gsdp_crore": {yr: round(v / 100, 1) for yr, v in g.items()},   # Lakh -> crore
            "nsdp_crore": {yr: round(v / 100, 1) for yr, v in n.items()},
            "per_capita_nsdp_inr": p,
            "population": pop,
        }

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT, "w") as f:
        json.dump({
            "_meta": {
                "source": "RBI Handbook of Statistics on Indian States 2024-25, Tables 19/21/23 (current prices, 2011-12 series).",
                "sourceUrl": SOURCE_URL,
                "units": "gsdp_crore & nsdp_crore in Rs crore (nominal); per_capita_nsdp_inr in Rs; population in persons (derived = NSDP/per-capita).",
                "note": "Covers 2011-12 to 2024-25. Population is derived, not a census count.",
            },
            "states": payload,
        }, f, indent=2, sort_keys=True)
        f.write("\n")

    print(f"wrote {OUT}  ({len(states)} states)")
    for s in ("Kerala", "Bihar", "Tamil Nadu", "Uttar Pradesh"):
        d = payload.get(s, {})
        g = d.get("gsdp_crore", {})
        pc_v = d.get("per_capita_nsdp_inr", {})
        pop = d.get("population", {})
        print(f"  {s}: GSDP 2023={g.get('2023')} cr  pc-NSDP 2023={pc_v.get('2023')}  pop 2023={pop.get('2023')}")


if __name__ == "__main__":
    main()
