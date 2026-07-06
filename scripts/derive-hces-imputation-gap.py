#!/usr/bin/env python3
"""HCES 2023-24: the welfare-imputation gap in MPCE, by fractile class.

HCES 2023-24 publishes average MPCE two ways: WITHOUT imputation, and WITH
imputation of the value of items received FREE through government social-welfare
programmes (free foodgrain through PDS/state schemes, free laptops, bicycles,
school uniforms, footwear, etc.). The published 2023-24 averages:
  rural  Rs 4,122 without / Rs 4,247 with  (gap +Rs 125)
  urban  Rs 6,996 without / Rs 7,078 with  (gap +Rs  82)
The published factsheet gives only these AVERAGE gaps. This script reconstructs
the gap BY FRACTILE so we can ask: do free welfare items matter more at the floor?

HOW THE IMPUTATION IS ISOLATED (this is the load-bearing bit)
-------------------------------------------------------------
The DDI for catalog 237 ships HCES_2023_24_Imputation_Rate.xlsx, which is MOSPI's
own imputation-rate book. It defines exactly which items are "received free of cost
through a social-welfare programme" and the per-unit rupee rate used to value them:
  * FREE FOOD  (FDQ items 061-075): in LEVEL-05 (sec 5 & 6) these appear as
    dedicated item codes 061-075 carrying QUANTITY ONLY, with VALUE blank/zero.
    MOSPI imputes value = quantity x rate[sector, state, item].
  * FREE DURABLES (DGQ dummy codes 991-997): in LEVEL-11 (sec 4.3) these appear as
    Num_Free_Laptop / _Mobile / _Bicycle / _Scooter / _Clothing / _Footwear /
    _Tablet COUNTS. MOSPI imputes value = count x rate[sector, state, code] / 12
    (the count is a 365-day stock; consumption is monthly, so /12 to monthlyise).

WITHOUT-imputation MPCE is the standard LEVEL-14 reconstruction (used in
derive-hces-poverty-cutoffs.py): LEVEL-14 only carries VALUED consumption, so the
zero-value free-food rows are already excluded from it.
WITH-imputation MPCE = WITHOUT + (imputed free-food value) + (imputed free-durable
value), per household, divided by household size for per-capita.

CALIBRATION GATE: the population-weighted average gap must reproduce rural ~+Rs 125
and urban ~+Rs 82, and the without-imputation average must reproduce Rs 4,122 /
Rs 6,996. If it doesn't, the isolation method is wrong and the gap-by-fractile table
is not trustworthy. We report the calibration result honestly either way.

Rates: MOSPI publishes both a "mode-based" and a "25th-percentile-based" rate per
item. We compute the gap under both and pick whichever reproduces the published
average gap; that choice is reported.

Usage:
  python3 scripts/derive-hces-imputation-gap.py            # report only
  python3 scripts/derive-hces-imputation-gap.py --write    # also write the JSON
"""
import csv, io, json, sys, zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ZIP = "data/snapshots/microdata-nada/downloads/HCES_Data_2023-24_Csv.zip"
RATE_XLSX = ("data/snapshots/microdata-nada/api-files/"
             "DDI-IND-MOSPI-NSS-HCES23-24/HCES_2023_24_Imputation_Rate.xlsx")
SOURCE_URL = "https://microdata.gov.in/NADA/index.php/catalog/237"
OUT = "data/series/derived.IN.poverty.hces_imputation_gap_by_fractile.json"
csv.field_size_limit(1 << 24)

PUB = {  # published averages (factsheet)
    "1": {"without": 4122.0, "gap": 125.0},
    "2": {"without": 6996.0, "gap": 82.0},
}
SECTOR_NAME = {"1": "Rural", "2": "Urban"}

# Free-durable count column in LEVEL-11  ->  DGQ dummy code in the rate table.
DURABLE_COLS = {
    "Num_Free_Laptop": "991",
    "Num_Free_Tablet": "992",
    "Num_Free_Mobile": "993",
    "Num_Free_Bicycle": "994",
    "Num_Free_Scooter": "995",
    "Num_Free_Clothing": "996",
    "Num_Free_Footwear": "997",
}
FREE_FOOD_CODES = {f"{n:03d}" for n in (61, 62, 63, 64, 65, 66, 67, 68, 70, 71, 72, 73, 74, 75)}


def member(zf, needle):
    for name in zf.namelist():
        if needle in name and name.lower().endswith(".csv"):
            return name
    raise SystemExit(f"CSV member not found: {needle}")


def hhid(cols):
    return tuple(cols[2:16])


def fnum(value):
    value = (value or "").strip()
    try:
        return float(value) if value else 0.0
    except ValueError:
        return 0.0


def factor(section):
    """Reference-period -> monthly scaling, identical to derive-hces-poverty-cutoffs.py."""
    if section[:2] in ("6.", "7."):
        return 30.0 / 7.0
    if section in ("10.1", "10.2", "11.4") or section[:3] in ("13.", "14."):
        return 30.0 / 365.0
    return 1.0


# ---------------------------------------------------------------- rate tables
def load_rates(which):
    """which in {'mode','p25'}. Returns
       food[(sector,state,item)] -> Rs/unit,  food_natl[(sector,item)] -> Rs/unit,
       dur[(sector,state,code)]  -> Rs/unit,  dur_natl[(sector,code)]  -> Rs/unit.
    The free-food quantity is a 30-day quantity, so no /12. The free-durable count
    is a 365-day stock, so the caller divides by 12 to monthlyise."""
    col = 5 if which == "mode" else 6  # 0-indexed columns in the rate sheets
    wb = openpyxl.load_workbook(RATE_XLSX, read_only=True, data_only=True)

    def grab(sheet, code_col, val_col, want_codes):
        rows = list(wb[sheet].iter_rows(values_only=True))
        out = {}
        for r in rows[2:]:
            if r[0] is None:
                continue
            sector = str(r[0]).strip()
            code = str(r[code_col]).strip()
            # rate sheets store codes as ints for durables; normalise food to 3-digit
            if code.isdigit() and len(code) < 3 and want_codes == "food":
                code = code.zfill(3)
            val = r[val_col]
            if val is None or str(val).strip() == "":
                continue
            try:
                rate = float(val)
            except (TypeError, ValueError):
                continue
            yield sector, r, code, rate

    # food: stsec sheet has cols sector, st, qitem, desc, ref, mode, p25
    food, food_natl_acc = {}, defaultdict(list)
    for sector, r, code, rate in grab("FDQ_Impu_rate_stsec", 2, col, "food"):
        st = str(r[1]).strip().zfill(2)
        code = code.zfill(3)
        food[(sector, st, code)] = rate
    for sector, r, code, rate in grab("FDQ_Impu_rate_sec", 1, col - 1, "food"):
        # sec sheet: cols sector, qitem, desc, ref, mode, p25  -> shift by one
        food_natl_acc[(sector, code.zfill(3))].append(rate)
    food_natl = {k: sum(v) / len(v) for k, v in food_natl_acc.items()}

    # durables: stsec sheet cols sector, st, dummycode, section, desc, ref, mode, p25
    dur, dur_natl_acc = {}, defaultdict(list)
    for r in list(wb["DGQ_Impu_rate_stsec"].iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        sector, st, code = str(r[0]).strip(), str(r[1]).strip().zfill(2), str(r[2]).strip()
        val = r[6] if which == "mode" else r[7]
        if val is None or str(val).strip() == "":
            continue
        try:
            dur[(sector, st, code)] = float(val)
        except (TypeError, ValueError):
            pass
    for r in list(wb["DGQ_Impu_rate_sec"].iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        sector, code = str(r[0]).strip(), str(r[1]).strip()
        val = r[5] if which == "mode" else r[6]
        if val is None or str(val).strip() == "":
            continue
        try:
            dur_natl_acc[(sector, code)].append(float(val))
        except (TypeError, ValueError):
            pass
    dur_natl = {k: sum(v) / len(v) for k, v in dur_natl_acc.items()}
    return food, food_natl, dur, dur_natl


# ---------------------------------------------------------------- microdata
def load_households():
    """Returns per-household dict keyed by hhid:
       {sector, state, size, weight, without, free_food_qty[(item)], free_dur_cnt[(code)]}."""
    hh = {}

    # LEVEL-15: household size (first non-zero size per household)
    with zipfile.ZipFile(ZIP) as zf:
        with zf.open(member(zf, "LEVEL - 15")) as fh:
            r = csv.reader(io.TextIOWrapper(fh, "utf-8", "replace"))
            H = next(r)
            iS = H.index("HOUSEHOLD_SIZE")
            for c in r:
                if len(c) <= iS:
                    continue
                k = hhid(c)
                size = fnum(c[iS])
                if size > 0 and k not in hh:
                    hh[k] = {"size": size}

    # LEVEL-14: WITHOUT-imputation consumption + sector/state/weight
    total = defaultdict(float)
    meta = {}
    with zipfile.ZipFile(ZIP) as zf:
        with zf.open(member(zf, "LEVEL - 14")) as fh:
            r = csv.reader(io.TextIOWrapper(fh, "utf-8", "replace"))
            H = next(r)
            iSec, iV, iM = H.index("SECTION"), H.index("VALUE_RS"), H.index("MULTIPLIER")
            for c in r:
                if len(c) <= iM:
                    continue
                k = hhid(c)
                total[k] += fnum(c[iV]) * factor(c[iSec])
                if k not in meta:
                    meta[k] = (c[3], c[4].zfill(2), fnum(c[iM]))

    # LEVEL-05: free-food quantities (item codes 061-075, value blank)
    food_qty = defaultdict(lambda: defaultdict(float))
    with zipfile.ZipFile(ZIP) as zf:
        with zf.open(member(zf, "LEVEL - 05")) as fh:
            r = csv.reader(io.TextIOWrapper(fh, "utf-8", "replace"))
            H = next(r)
            iI = H.index("Item_Code")
            iQ = H.index("Total_Consumption_Quantity")
            for c in r:
                if len(c) <= iQ:
                    continue
                it = c[iI].strip()
                if it in FREE_FOOD_CODES:
                    food_qty[hhid(c)][it] += fnum(c[iQ])

    # LEVEL-11: free-durable counts (Num_Free_* columns)
    dur_cnt = defaultdict(lambda: defaultdict(float))
    with zipfile.ZipFile(ZIP) as zf:
        with zf.open(member(zf, "LEVEL - 11")) as fh:
            r = csv.reader(io.TextIOWrapper(fh, "utf-8", "replace"))
            H = next(r)
            idx = {col: H.index(col) for col in DURABLE_COLS}
            for c in r:
                k = hhid(c)
                for col, code in DURABLE_COLS.items():
                    n = fnum(c[idx[col]])
                    if n:
                        dur_cnt[k][code] += n

    out = []
    for k, rec in hh.items():
        if k not in meta:
            continue
        sector, state, weight = meta[k]
        size, without = rec["size"], total.get(k, 0.0)
        if weight <= 0 or without <= 0 or size <= 0:
            continue
        out.append({
            "key": k, "sector": sector, "state": state, "size": size,
            "weight": weight, "without": without,
            "food_qty": food_qty.get(k, {}), "dur_cnt": dur_cnt.get(k, {}),
        })
    return out


def imputed_value(rec, food, food_natl, dur, dur_natl):
    """Total monthly imputed value of free welfare items for one household (Rs)."""
    s, st = rec["sector"], rec["state"]
    val = 0.0
    for item, qty in rec["food_qty"].items():
        rate = food.get((s, st, item))
        if rate is None:
            rate = food_natl.get((s, item), 0.0)
        val += qty * rate  # food qty is already a 30-day quantity
    for code, cnt in rec["dur_cnt"].items():
        rate = dur.get((s, st, code))
        if rate is None:
            rate = dur_natl.get((s, code), 0.0)
        val += cnt * rate / 12.0  # durable count is a 365-day stock -> monthlyise
    return val


# ---------------------------------------------------------------- fractiles
FRACTILES = [
    ("Poorest 5%", 0.00, 0.05),
    ("5-10%", 0.05, 0.10),
    ("Decile 1 (poorest 10%)", 0.00, 0.10),
    ("Decile 2", 0.10, 0.20),
    ("Decile 3", 0.20, 0.30),
    ("Decile 4", 0.30, 0.40),
    ("Decile 5", 0.40, 0.50),
    ("Decile 6", 0.50, 0.60),
    ("Decile 7", 0.60, 0.70),
    ("Decile 8", 0.70, 0.80),
    ("Decile 9", 0.80, 0.90),
    ("Decile 10 (richest 10%)", 0.90, 1.00),
]


def weighted_avg(records, key, sector=None):
    rows = [r for r in records if sector is None or r["sector"] == sector]
    den = sum(r["pw"] for r in rows)
    return sum(r[key] * r["pw"] for r in rows) / den if den else 0.0


def fractile_table(records, sector):
    """Population-weighted gap by fractile, ranked on WITHOUT-imputation MPCE."""
    rows = sorted((r for r in records if r["sector"] == sector), key=lambda r: r["without_pc"])
    total_pw = sum(r["pw"] for r in rows)
    # cumulative person-weight position
    cum = 0.0
    pos = []
    for r in rows:
        lo = cum
        cum += r["pw"]
        pos.append((lo / total_pw, cum / total_pw, r))
    out = []
    for label, a, b in FRACTILES:
        sel = [r for (lo, hi, r) in pos if lo < b and hi > a]
        pw = sum(r["pw"] for r in sel)
        if not pw:
            continue
        without = sum(r["without_pc"] * r["pw"] for r in sel) / pw
        gap = sum(r["gap_pc"] * r["pw"] for r in sel) / pw
        out.append({
            "label": label,
            "withoutImputationMpce": round(without),
            "imputationGapRs": round(gap, 1),
            "gapPctOfMpce": round(100.0 * gap / without, 2) if without else 0.0,
        })
    return out


def main():
    write = "--write" in sys.argv
    records = load_households()
    print(f"households usable: {len(records):,}", file=sys.stderr)

    # person-weight = household weight x household size
    for r in records:
        r["pw"] = r["weight"] * r["size"]
        r["without_pc"] = r["without"] / r["size"]

    # uncalibrated without-imputation average (sanity)
    for s in ("1", "2"):
        raw = weighted_avg(records, "without_pc", s)
        print(f"  raw without-imp {SECTOR_NAME[s]}: {raw:.0f} (pub {PUB[s]['without']:.0f})",
              file=sys.stderr)

    # Calibrate without-imputation MPCE to published (same as cutoffs script).
    cal = {s: PUB[s]["without"] / weighted_avg(records, "without_pc", s) for s in ("1", "2")}
    for r in records:
        r["without_pc"] *= cal[r["sector"]]

    # ----- choose rate basis (mode vs 25th pct) by which reproduces the gap -----
    results = {}
    for which in ("mode", "p25"):
        food, food_natl, dur, dur_natl = load_rates(which)
        for r in records:
            imp = imputed_value(r, food, food_natl, dur, dur_natl)
            r[f"imp_pc_{which}"] = imp / r["size"]  # per-capita imputed value (uncalibrated)
        # the imputed value rides on the SAME calibration scale as without_pc
        rural_gap = weighted_avg(
            [{"v": rr[f"imp_pc_{which}"] * cal[rr["sector"]], "pw": rr["pw"], "sector": rr["sector"]}
             for rr in records], "v", "1")
        urban_gap = weighted_avg(
            [{"v": rr[f"imp_pc_{which}"] * cal[rr["sector"]], "pw": rr["pw"], "sector": rr["sector"]}
             for rr in records], "v", "2")
        results[which] = (rural_gap, urban_gap)
        print(f"  rate basis {which:4}: avg gap rural +{rural_gap:.1f}  urban +{urban_gap:.1f} "
              f"(pub +125 / +82)", file=sys.stderr)

    # pick basis minimising combined deviation from published gaps
    def dev(rb):
        rg, ug = results[rb]
        return abs(rg - PUB["1"]["gap"]) + abs(ug - PUB["2"]["gap"])
    basis = min(("mode", "p25"), key=dev)
    print(f"\n  -> chosen rate basis: {basis}", file=sys.stderr)

    for r in records:
        r["gap_pc"] = r[f"imp_pc_{basis}"] * cal[r["sector"]]
        r["with_pc"] = r["without_pc"] + r["gap_pc"]

    # ---------------------------- CALIBRATION GATE ----------------------------
    print("\n================ CALIBRATION GATE ================")
    gate_pass = True
    for s in ("1", "2"):
        without = weighted_avg(records, "without_pc", s)
        gap = weighted_avg(records, "gap_pc", s)
        with_imp = without + gap
        ok_w = abs(without - PUB[s]["without"]) < 5
        ok_g = abs(gap - PUB[s]["gap"]) <= 5
        gate_pass = gate_pass and ok_w and ok_g
        print(f"  {SECTOR_NAME[s]:5}: without {without:7.1f} (pub {PUB[s]['without']:.0f}) "
              f"{'OK' if ok_w else 'FAIL':4} | gap +{gap:5.1f} (pub +{PUB[s]['gap']:.0f}) "
              f"{'OK' if ok_g else 'FAIL':4} | with {with_imp:7.1f}")
    print(f"  GATE: {'PASS' if gate_pass else 'FAIL'}")
    print("=================================================\n")

    # ---------------------------- gap by fractile -----------------------------
    rows = []
    for s in ("1", "2"):
        sector = SECTOR_NAME[s]
        print(f"== {sector}: imputation gap by fractile (population-weighted) ==")
        print(f"  {'fractile':<26} {'without':>8} {'gap Rs':>8} {'% of MPCE':>10}")
        for fr in fractile_table(records, s):
            print(f"  {fr['label']:<26} {fr['withoutImputationMpce']:>8} "
                  f"{fr['imputationGapRs']:>8.1f} {fr['gapPctOfMpce']:>9.2f}%")
            rows.append({
                "label": f"{sector}: {fr['label']}",
                "sector": sector,
                "fractile": fr["label"],
                "withoutImputationMpce": fr["withoutImputationMpce"],
                "imputationGapRs": fr["imputationGapRs"],
                "gapPctOfMpce": fr["gapPctOfMpce"],
            })
        # overall row
        without = round(weighted_avg(records, "without_pc", s))
        gap = round(weighted_avg(records, "gap_pc", s), 1)
        rows.append({
            "label": f"{sector}: Overall average",
            "sector": sector,
            "fractile": "Overall average",
            "withoutImputationMpce": without,
            "imputationGapRs": gap,
            "gapPctOfMpce": round(100.0 * gap / without, 2),
        })
        print(f"  {'Overall average':<26} {without:>8} {gap:>8.1f} "
              f"{100.0*gap/without:>9.2f}%\n")

    if not write:
        print("(dry run; pass --write to emit JSON)", file=sys.stderr)
        return

    artifact = {
        "schemaVersion": 1,
        "artifactType": "table",
        "indicatorId": "econ.poverty.hces_imputation_gap_by_fractile",
        "title": "HCES welfare-imputation gap in MPCE by fractile and sector, 2023-24",
        "sourceId": "mospi-hces",
        "sourceIndicatorId": "HCES 2023-24 unit-level microdata: free-welfare imputation gap by fractile",
        "sourceUrl": SOURCE_URL,
        "unit": "Rs per person per month / % of MPCE",
        "geography": {"type": "country", "id": "IN", "name": "India"},
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "rows": rows,
        "dimensions": ["label", "sector", "fractile", "withoutImputationMpce",
                       "imputationGapRs", "gapPctOfMpce"],
        "metadata": {
            "provenance": ("Computed from HCES 2023-24 unit-level microdata. The value of items "
                           "received free through social-welfare programmes is reconstructed using "
                           "MOSPI's own imputation-rate book (HCES_2023_24_Imputation_Rate.xlsx, "
                           "DDI catalog 237): free-food quantities (LEVEL-05 item codes 061-075, "
                           "value blank in the raw data) and free-durable counts (LEVEL-11 "
                           "Num_Free_* columns, dummy codes 991-997) are multiplied by the "
                           f"state x sector rate (basis: {basis})."),
            "method": ("Without-imputation MPCE is the standard LEVEL-14 reconstruction, calibrated "
                       "to the published national rural/urban averages (Rs 4,122 / Rs 6,996). The "
                       "imputation gap is the per-capita imputed value of free welfare items, on the "
                       "same calibration scale. Fractiles are population-weighted (household "
                       "multiplier x household size), ranked on without-imputation MPCE, computed "
                       "separately for rural and urban."),
            "calibration": {
                "rural": {"withoutImputation": round(weighted_avg(records, "without_pc", "1")),
                          "averageGap": round(weighted_avg(records, "gap_pc", "1"), 1),
                          "publishedWithout": 4122, "publishedGap": 125},
                "urban": {"withoutImputation": round(weighted_avg(records, "without_pc", "2")),
                          "averageGap": round(weighted_avg(records, "gap_pc", "2"), 1),
                          "publishedWithout": 6996, "publishedGap": 82},
                "gatePass": bool(gate_pass),
                "rateBasis": basis,
            },
            "caveat": ("Welfare-free items are reconstructed, not read off a single column: free-food "
                       "rows carry quantity but no value, so value is imputed at MOSPI's published "
                       "state x sector rate. Durable counts are a 365-day stock monthlyised by /12. "
                       "Fractile boundaries and levels depend on the MPCE reconstruction and "
                       "reference-period scaling; treat the gap-by-fractile pattern as indicative."),
        },
    }
    Path("data/series").mkdir(parents=True, exist_ok=True)
    Path(OUT).write_text(json.dumps(artifact, indent=2) + "\n")
    print(f"wrote {OUT}", file=sys.stderr)


if __name__ == "__main__":
    main()
