#!/usr/bin/env python3
"""Ingest RBI Bulletin Table 36, outward remittances under LRS.

The RBI workbook is a monthly administrative table in US dollar crore. It has
ten top-level purposes and, from March 2026, five nested travel purposes. This
script preserves every published field, converts values to US dollar billion,
builds fiscal-year series, and fails loudly if the workbook shape or notes move.

Important arithmetic rule: the travel subcategories are children of Travel.
They must never be added to the ten top-level purposes.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import os
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DOWNLOADS = Path.home() / "Downloads"
OUT = Path("data/series")
SNAP = Path("data/snapshots/rbi-lrs")
CATALOG = Path("data/catalog")
SOURCE_URL = "https://www.rbi.org.in/scripts/BS_ViewBulletin.aspx?Id=24281"
FETCHED_AT = dt.datetime.now(dt.timezone.utc).isoformat()

TITLE = "Outward Remittances Under The Liberalised Remittance Scheme For Resident Individuals - Monthly"
UNIT_LABEL = "(US$ Crores)"
SOURCE_NOTE = "Source : Reserve Bank of India"
OTHERS_NOTE = "** Include items such as subscription to journals, maintenance of investment abroad, student loan repayments and credit card payments."

FIELDS = [
    ("deposit", "1.1. Deposit", "Deposit", False),
    ("property", "1.2. Purchase of immovable property", "Purchase of immovable property", False),
    ("investment", "1.3. Investment in equity/debt", "Investment in equity/debt", False),
    ("gift", "1.4. Gift", "Gift", False),
    ("donations", "1.5. Donations", "Donations", False),
    ("travel", "1.6.Travels", "Travel", False),
    ("travel_business", "1.6.1  Business travel", "Business travel", True),
    ("travel_pilgrimage", "1.6.2  Travel for pilgrimage", "Travel for pilgrimage", True),
    ("travel_medical", "1.6.3  Travel for medical treatment", "Travel for medical treatment", True),
    ("travel_education", "1.6.4  Travel for education  (including fees, hostel expenses etc.)", "Travel for education", True),
    ("travel_other", "1.6.5  Other travel (including holiday trips and payments for settling international credit cards transactions)", "Other travel", True),
    ("relatives", "1.7.Maintenance of close relatives", "Maintenance of close relatives", False),
    ("medical", "1.8.Medical Treatment", "Medical treatment", False),
    ("studies", "1.9.Studies Abroad", "Studies abroad", False),
    ("others", "1.10. Others**", "Others", False),
]

TOP_LEVEL = [key for key, _, _, nested in FIELDS if not nested]
TRAVEL_CHILDREN = [key for key, _, _, nested in FIELDS if nested]
ASSET_BUILDING = ["deposit", "property", "investment"]

EXPECTED_NULLS = {
    "deposit": 0,
    "property": 0,
    "investment": 0,
    "gift": 0,
    "donations": 3,
    "travel": 12,
    "travel_business": 215,
    "travel_pilgrimage": 215,
    "travel_medical": 215,
    "travel_education": 215,
    "travel_other": 215,
    "relatives": 12,
    "medical": 12,
    "studies": 12,
    "others": 0,
}


def source_path() -> Path:
    explicit = os.environ.get("RBI_LRS_XLSX")
    if explicit:
        path = Path(explicit).expanduser()
        if not path.exists():
            raise FileNotFoundError(path)
        return path
    matches = list(DOWNLOADS.glob("RBIB Table No. 36 _ Outward Remittances under the Liberalised Remittance Scheme for Resident Individuals*.xlsx"))
    if not matches:
        raise FileNotFoundError("No RBI LRS Table 36 workbook found in ~/Downloads")
    return max(matches, key=lambda path: path.stat().st_mtime)


def assert_equal(actual, expected, label: str) -> None:
    if actual != expected:
        raise ValueError(f"{label} changed: expected {expected!r}, got {actual!r}")


def month_key(value: dt.datetime) -> str:
    return value.strftime("%Y-%m")


def fiscal_year_start(value: dt.datetime) -> int:
    return value.year if value.month >= 4 else value.year - 1


def fiscal_end_date(start_year: int) -> str:
    return f"{start_year + 1}-03-31"


def artifact_base(indicator_id: str, title: str, frequency: str, unit: str, metadata: dict) -> dict:
    return {
        "schemaVersion": 1,
        "artifactType": "series",
        "indicatorId": indicator_id,
        "title": title,
        "sourceId": "rbi-lrs",
        "sourceIndicatorId": "RBI Bulletin Table 36",
        "sourceUrl": SOURCE_URL,
        "unit": unit,
        "frequency": frequency,
        "geography": {"type": "country", "id": "IN", "name": "India"},
        "dimensions": [],
        "fetchedAt": FETCHED_AT,
        "metadata": metadata,
    }


def write_series(filename: str, indicator_id: str, title: str, frequency: str, observations: list[dict], metadata: dict) -> dict:
    artifact = artifact_base(indicator_id, title, frequency, "US$ billion", metadata)
    artifact["observations"] = observations
    path = OUT / f"{filename}.json"
    path.write_text(json.dumps(artifact, indent=2) + "\n")
    return {"indicatorId": indicator_id, "path": str(path), "observations": len(observations)}


def write_table(filename: str, indicator_id: str, title: str, unit: str, rows: list[dict], metadata: dict) -> dict:
    artifact = artifact_base(indicator_id, title, "snapshot", unit, metadata)
    artifact["artifactType"] = "table"
    artifact.pop("frequency")
    artifact["rows"] = rows
    path = OUT / f"{filename}.json"
    path.write_text(json.dumps(artifact, indent=2) + "\n")
    return {"indicatorId": indicator_id, "path": str(path), "rows": len(rows)}


def main() -> None:
    src = source_path()
    raw = src.read_bytes()
    raw_hash = hashlib.sha256(raw).hexdigest()
    SNAP.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    CATALOG.mkdir(parents=True, exist_ok=True)
    snapshot_path = SNAP / f"rbi-lrs-table-36.{raw_hash[:12]}.xlsx"
    shutil.copyfile(src, snapshot_path)

    workbook = load_workbook(src, data_only=False, read_only=False)
    assert_equal(len(workbook.worksheets), 1, "worksheet count")
    sheet = workbook.active
    assert_equal(sheet.sheet_state, "visible", "worksheet visibility")
    assert_equal(sheet.calculate_dimension(), "B2:Q228", "used range")
    assert_equal(sheet["B2"].value, TITLE, "table title")
    assert_equal(sheet["B4"].value, UNIT_LABEL, "unit label")
    assert_equal(sheet["B226"].value, SOURCE_NOTE, "source footer")
    assert_equal(sheet["B228"].value, OTHERS_NOTE, "methodology footer")
    assert_equal([sheet.cell(6, col).value for col in range(3, 18)], [field[1] for field in FIELDS], "purpose headers")
    assert_equal([sheet.cell(7, col).value for col in range(3, 18)], list(range(1, 16)), "purpose column numbers")

    formulas = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]
    comments = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.comment]
    hyperlinks = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.hyperlink]
    assert_equal(formulas, [], "formula cells")
    assert_equal(comments, [], "comment cells")
    assert_equal(hyperlinks, [], "hyperlink cells")
    assert_equal(list(workbook.defined_names.values()), [], "defined names")

    records = []
    for excel_row in range(8, 225):
        date = sheet.cell(excel_row, 2).value
        if not isinstance(date, dt.datetime):
            raise ValueError(f"B{excel_row} is not a date: {date!r}")
        values = {}
        for col, (key, _, _, _) in enumerate(FIELDS, start=3):
            raw_value = sheet.cell(excel_row, col).value
            if raw_value == "-":
                values[key] = None
            elif isinstance(raw_value, (int, float)):
                if raw_value < 0 or raw_value > 10_000:
                    raise ValueError(f"{sheet.cell(excel_row, col).coordinate} outside plausible US$ crore range: {raw_value}")
                values[key] = float(raw_value) / 100.0  # US$ crore to US$ billion
            else:
                raise ValueError(f"Unexpected value at {sheet.cell(excel_row, col).coordinate}: {raw_value!r}")
        records.append({"date": date, "values": values, "excelRow": excel_row})

    assert_equal(len(records), 217, "monthly row count")
    records.sort(key=lambda row: row["date"])
    assert_equal(month_key(records[0]["date"]), "2008-04", "first month")
    assert_equal(month_key(records[-1]["date"]), "2026-04", "last month")
    for previous, current in zip(records, records[1:]):
        expected_year = previous["date"].year + (1 if previous["date"].month == 12 else 0)
        expected_month = 1 if previous["date"].month == 12 else previous["date"].month + 1
        if (current["date"].year, current["date"].month) != (expected_year, expected_month):
            raise ValueError(f"Non-consecutive months: {previous['date']} then {current['date']}")

    null_counts = Counter()
    for record in records:
        for key, value in record["values"].items():
            if value is None:
                null_counts[key] += 1
    observed_nulls = {key: null_counts[key] for key, _, _, _ in FIELDS}
    assert_equal(observed_nulls, EXPECTED_NULLS, "dash/null pattern")

    travel_split_months = []
    for record in records:
        children = [record["values"][key] for key in TRAVEL_CHILDREN]
        if any(value is not None for value in children):
            if not all(value is not None for value in children):
                raise ValueError(f"Partial travel split in {month_key(record['date'])}")
            travel = record["values"]["travel"]
            child_sum = sum(children)
            if travel is None or abs(travel - child_sum) > 1e-9:
                raise ValueError(f"Travel children do not sum to Travel in {month_key(record['date'])}: {child_sum} vs {travel}")
            travel_split_months.append(month_key(record["date"]))
    assert_equal(travel_split_months, ["2026-03", "2026-04"], "travel split coverage")

    common_metadata = {
        "rbiTable": "RBI Bulletin Table 36",
        "rawWorkbook": src.name,
        "rawHash": raw_hash,
        "snapshotPath": str(snapshot_path),
        "rawUnit": "US$ crore",
        "conversion": "raw workbook value divided by 100 to obtain US$ billion",
        "kind": "administrative measurement of gross outward flows reported under LRS",
        "geographyDefinition": "resident individuals in India; no destination geography",
        "footerNotes": [SOURCE_NOTE, OTHERS_NOTE],
        "dashPolicy": "Literal dashes are preserved as null. Sparse top-level dashes contribute no published amount to the reconciled headline sum; historic travel-child dashes mean the breakdown was not published.",
        "coverageBreak": "From May 26, 2015, RBI subsumed several current-account facilities, including private and business travel, under the common LRS limit. Pre-break totals are not directly comparable with later totals.",
        "othersInterpretation": "The workbook footer makes Others a heterogeneous residual spanning journals, maintenance of investments abroad, student-loan repayments and credit-card payments. It must not be treated as a single behaviour.",
        "travelClassificationNote": "The detailed travel children are published only for March and April 2026. RBI's June 2026 bulletin distinguishes travel for education or medical care from remotely purchased education or medical services in the separate Studies Abroad and Medical Treatment rows.",
    }

    manifest_entries = []
    labels = {key: label for key, _, label, _ in FIELDS}
    for key, _, label, nested in FIELDS:
        observations = [
            {"date": month_key(record["date"]), "value": round(record["values"][key], 9)}
            for record in records
            if record["values"][key] is not None
        ]
        manifest_entries.append(write_series(
            f"rbi-lrs.IN.extfin.lrs.{key}.monthly",
            f"extfin.lrs.{key}.monthly",
            f"LRS outward remittances: {label}, monthly",
            "monthly",
            observations,
            {**common_metadata, "purpose": label, "nestedUnderTravel": nested, "method": "Monthly RBI value converted from US$ crore to US$ billion; dashes omitted from observations."},
        ))

    # Headline monthly total. Nested travel fields are explicitly excluded.
    monthly_total = []
    for record in records:
        value = sum(record["values"][key] or 0.0 for key in TOP_LEVEL)
        monthly_total.append({"date": month_key(record["date"]), "value": round(value, 9)})
    manifest_entries.append(write_series(
        "rbi-lrs.IN.extfin.lrs.total.monthly",
        "extfin.lrs.total.monthly",
        "Total outward remittances under LRS, monthly",
        "monthly",
        monthly_total,
        {**common_metadata, "method": "Sum of the ten top-level purpose columns only. Travel children are excluded to prevent double counting."},
    ))

    by_fy = defaultdict(list)
    for record in records:
        by_fy[fiscal_year_start(record["date"])].append(record)
    complete_fys = {year: rows for year, rows in by_fy.items() if len(rows) == 12 and year >= 2009}
    assert_equal(sorted(complete_fys), list(range(2009, 2026)), "complete fiscal years")

    annual_values = defaultdict(dict)
    for year, rows in sorted(complete_fys.items()):
        for key in TOP_LEVEL:
            annual_values[key][year] = sum(row["values"][key] or 0.0 for row in rows)
        annual_values["total"][year] = sum(annual_values[key][year] for key in TOP_LEVEL)
        annual_values["asset_building"][year] = sum(annual_values[key][year] for key in ASSET_BUILDING)

    # Published RBI FY2025-26 totals are the independent table-level arithmetic check.
    published_2025_26 = {
        "total": 28.97936,
        "deposit": 0.75308,
        "property": 0.52870,
        "investment": 2.65201,
        "gift": 2.59481,
        "donations": 0.01212,
        "travel": 16.43799,
        "relatives": 3.53904,
        "medical": 0.05855,
        "studies": 2.30974,
        "others": 0.09330,
    }
    reconciliation = {}
    for key, published in published_2025_26.items():
        workbook_value = annual_values[key][2025]
        gap = workbook_value - published
        if abs(gap) > 0.00002:
            raise ValueError(f"FY2025-26 {key} does not reconcile with RBI Bulletin: {workbook_value} vs {published}")
        reconciliation[key] = {"workbookUSDBillion": round(workbook_value, 9), "publishedUSDBillion": published, "gapUSDBillion": round(gap, 9)}

    # Preserve RBI's displayed annual total separately. This value is not
    # derived from the workbook and powers a publication-level mirror check.
    manifest_entries.append(write_series(
        "rbi-bulletin-lrs.IN.extfin.lrs.total.fy.published",
        "extfin.lrs.total.fy.published",
        "RBI-published LRS outward remittances total, FY2025-26",
        "annual",
        [{"date": "2026-03-31", "value": published_2025_26["total"]}],
        {
            **common_metadata,
            "method": "Annual total as displayed in RBI Bulletin Table 36 for FY2025-26; not recomputed from the downloaded monthly workbook.",
            "reconciliationRole": "Publication-level mirror for the workbook-derived fiscal-year total.",
        },
    ))

    annual_keys = TOP_LEVEL + ["asset_building", "total"]
    for key in annual_keys:
        label = "Clearly asset-building purposes" if key == "asset_building" else "Total" if key == "total" else labels[key]
        observations = [
            {"date": fiscal_end_date(year), "value": round(annual_values[key][year], 9)}
            for year in sorted(complete_fys)
        ]
        method = "Fiscal-year sum of monthly values."
        if key == "asset_building":
            method += " Derived as Deposit + Purchase of immovable property + Investment in equity/debt; the label is descriptive, not an exhaustive legal capital-account classification."
        if key == "total":
            method += " Sum of ten top-level purposes; nested travel children excluded."
        manifest_entries.append(write_series(
            f"rbi-lrs.IN.extfin.lrs.{key}.fy",
            f"extfin.lrs.{key}.fy",
            f"LRS outward remittances: {label}, fiscal year",
            "annual",
            observations,
            {**common_metadata, "method": method, "fiscalYear": "April to March", "reconciliationFY2025_26": reconciliation.get(key)},
        ))

    latest_year = 2025
    latest_total = annual_values["total"][latest_year]
    latest_rows = []
    for key in sorted(TOP_LEVEL, key=lambda item: annual_values[item][latest_year], reverse=True):
        latest_rows.append({
            "label": labels[key],
            "value": round(annual_values[key][latest_year], 9),
            "sharePercent": round(annual_values[key][latest_year] / latest_total * 100, 3),
            "group": "clearly asset-building" if key in ASSET_BUILDING else "other reported purpose",
        })
    manifest_entries.append(write_table(
        "rbi-lrs.IN.extfin.lrs.purpose.latest",
        "extfin.lrs.purpose.latest",
        "LRS outward remittances by purpose, FY2025-26",
        "US$ billion",
        latest_rows,
        {**common_metadata, "method": "Fiscal-year sum for FY2025-26, ranked by value; shares use the ten-purpose headline total.", "period": "FY2025-26"},
    ))

    # Two-point purpose series let rankedChange compare the FY2023-24 peak with FY2025-26.
    for key in TOP_LEVEL:
        observations = [
            {"date": "2024-03-31", "value": round(annual_values[key][2023], 9)},
            {"date": "2026-03-31", "value": round(annual_values[key][2025], 9)},
        ]
        manifest_entries.append(write_series(
            f"rbi-lrs.IN.extfin.lrs.{key}.peak_change",
            f"extfin.lrs.{key}.peak_change",
            f"LRS {labels[key]}: FY2023-24 peak versus FY2025-26",
            "annual",
            observations,
            {**common_metadata, "method": "Two-point extract of fiscal-year sums, comparing FY2023-24 with FY2025-26.", "purpose": labels[key]},
        ))

    split_records = [record for record in records if month_key(record["date"]) in travel_split_months]
    split_total = sum(record["values"]["travel"] or 0.0 for record in split_records)
    split_rows = []
    for key in sorted(TRAVEL_CHILDREN, key=lambda item: sum(record["values"][item] or 0.0 for record in split_records), reverse=True):
        value = sum(record["values"][key] or 0.0 for record in split_records)
        split_rows.append({"label": labels[key], "value": round(value / split_total * 100, 3), "usdBillion": round(value, 9)})
    manifest_entries.append(write_table(
        "rbi-lrs.IN.extfin.lrs.travel_split.mar_apr_2026",
        "extfin.lrs.travel_split.mar_apr_2026",
        "What RBI's Travel category contained in March-April 2026",
        "percent of Travel",
        split_rows,
        {**common_metadata, "share": True, "period": "March-April 2026 only", "method": "Sum each published travel child over March and April 2026, then divide by the two-month Travel total. Do not extrapolate these shares to other months."},
    ))

    audit = {
        "sourcePath": src.name,
        "snapshotPath": str(snapshot_path),
        "sha256": raw_hash,
        "workbookCreated": workbook.properties.created.isoformat() if workbook.properties.created else None,
        "sheet": sheet.title,
        "usedRange": sheet.calculate_dimension(),
        "monthlyRows": len(records),
        "dateRange": [month_key(records[0]["date"]), month_key(records[-1]["date"])],
        "fields": [{"key": key, "header": header, "label": label, "nestedUnderTravel": nested, "nulls": EXPECTED_NULLS[key]} for key, header, label, nested in FIELDS],
        "travelSplitMonths": travel_split_months,
        "footerNotes": [SOURCE_NOTE, OTHERS_NOTE],
        "reconciliationFY2025_26": reconciliation,
        "doubleCountGuard": "Headline totals use ten top-level purposes only; five travel children are excluded.",
        "interpretation": {
            "others": common_metadata["othersInterpretation"],
            "coverageBreak": common_metadata["coverageBreak"],
            "dashPolicy": common_metadata["dashPolicy"],
        },
    }
    audit_path = CATALOG / "rbi-lrs-workbook-audit.json"
    audit_path.write_text(json.dumps(audit, indent=2) + "\n")
    manifest_entries = [
        {
            "status": "ready",
            "indicatorId": entry["indicatorId"],
            "sourceIndicatorId": "RBI Bulletin Table 36",
            "fetchedAt": FETCHED_AT,
            "artifact": entry["path"],
            **({"observations": entry["observations"]} if "observations" in entry else {"rows": entry["rows"]}),
            "snapshot": str(snapshot_path),
            "rawHash": raw_hash,
        }
        for entry in manifest_entries
    ]
    manifest_path = CATALOG / "rbi-lrs-manifest.json"
    manifest_path.write_text(json.dumps(manifest_entries, indent=2) + "\n")

    print(f"audited {src}")
    print(f"snapshot {snapshot_path} sha256={raw_hash}")
    print(f"wrote {len(manifest_entries)} artifacts and {audit_path}")
    print(f"FY2025-26 total: {annual_values['total'][2025]:.9f} US$ billion")


if __name__ == "__main__":
    main()
